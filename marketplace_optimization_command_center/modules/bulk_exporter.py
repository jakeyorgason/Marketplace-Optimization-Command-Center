from __future__ import annotations

from pathlib import Path
from datetime import datetime
import re
import zipfile
import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    from modules.config import EXPORTS_DIR, APP_DIR
    from modules.data_loader import slugify, list_reports, load_processed_report
except Exception:  # pragma: no cover
    from .config import EXPORTS_DIR, APP_DIR
    from .data_loader import slugify, list_reports, load_processed_report

BULK_EXPORTER_VERSION = "2026-06-30-valid-id-resolver-v6"

TEMPLATE_PATH = APP_DIR / "templates" / "PxP_SP_Bulk_Upload_Jun25_2026_FIXED_v2.xlsx"
DEFAULT_COLUMNS = [
    "Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID", "Product Targeting ID",
    "Campaign Name", "Ad Group Name", "State", "Keyword Text", "Match Type", "Bid", "Budget", "Daily Budget",
    "Placement Type", "Placement %", "Portfolio ID", "Product Targeting Expression",
]
PRODUCT_BY_AD_TYPE = {"SP": "Sponsored Products", "SB": "Sponsored Brands", "SD": "Sponsored Display"}
BRAND_REASON_PREFIX = "brand_safety"


def _out_dir(client_name: str) -> Path:
    out = EXPORTS_DIR / slugify(client_name)
    out.mkdir(parents=True, exist_ok=True)
    return out


def _safe_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "0.0"}:
        return ""
    return text


def _safe_float(value, default: float = 0.0) -> float:
    try:
        if pd.isna(value):
            return default
        text = str(value).replace("$", "").replace(",", "").strip()
        if text == "":
            return default
        return float(text)
    except Exception:
        return default


def _norm(value) -> str:
    return re.sub(r"\s+", " ", _safe_text(value).lower()).strip()


def _norm_id(value) -> str:
    text = _safe_text(value)
    if text.endswith(".0") and text.replace(".0", "").isdigit():
        return text[:-2]
    return text


def _raw(row: pd.Series, *names: str) -> str:
    for name in names:
        candidates = [
            name,
            "raw__" + re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_"),
            re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_"),
        ]
        for c in candidates:
            if c in row.index:
                val = _safe_text(row.get(c))
                if val:
                    return _norm_id(val)
    return ""


def _template_headers() -> list[str]:
    if TEMPLATE_PATH.exists() and load_workbook:
        try:
            wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=False)
            ws = wb["Output"] if "Output" in wb.sheetnames else wb[wb.sheetnames[0]]
            headers = [_safe_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [h for h in headers if h]
            if headers:
                return headers
        except Exception:
            pass
    return DEFAULT_COLUMNS.copy()


def _target_for_negative(row: pd.Series) -> str:
    return _safe_text(row.get("target")) or _raw(row, "customer search term", "search term", "keyword text", "targeting")


def _asin_from_text(value: str) -> str:
    text = _safe_text(value).upper()
    m = re.search(r"\bB0[A-Z0-9]{8}\b", text)
    return m.group(0) if m else ""


def _product_expression_from_target(value: str) -> str:
    text = _safe_text(value)
    if not text:
        return ""
    if text.lower().startswith("asin=") or text.lower().startswith("asin=") or "asin=\"" in text.lower():
        return text
    asin = _asin_from_text(text)
    if asin:
        return f'asin="{asin.lower()}"'
    return text


def _is_product_target(row: pd.Series) -> bool:
    target = _safe_text(row.get("target"))
    expr = _raw(row, "product targeting expression", "resolved product targeting expression", "resolved product targeting expression (informational only)")
    ptid = _raw(row, "product targeting id")
    entity = (_raw(row, "entity") or _safe_text(row.get("entity"))).lower()
    return bool(expr or ptid or "product targeting" in entity or _asin_from_text(target))


def _is_brand_reason(row: pd.Series) -> bool:
    return _safe_text(row.get("reason_code")).startswith(BRAND_REASON_PREFIX)


def _product(ad_type: str, row: pd.Series | None = None) -> str:
    ad_type = (ad_type or _safe_text(row.get("ad_type") if row is not None else "")).upper()
    return PRODUCT_BY_AD_TYPE.get(ad_type, "Sponsored Products")


def _latest_report_group(client_name: str) -> pd.DataFrame:
    reports = list_reports(client_name)
    if reports is None or reports.empty:
        return pd.DataFrame()
    reports = reports.copy()
    reports["uploaded_at"] = pd.to_datetime(reports["uploaded_at"], errors="coerce")
    latest = reports.sort_values("uploaded_at", ascending=False).iloc[0]
    if "hash" in reports.columns and pd.notna(latest.get("hash")):
        same = reports[reports["hash"].astype(str) == str(latest["hash"])]
        if not same.empty:
            return same
    return reports.head(20)


def _load_lookup_df(client_name: str | None) -> pd.DataFrame:
    if not client_name:
        return pd.DataFrame()
    group = _latest_report_group(client_name)
    frames = []
    for _, rec in group.iterrows():
        try:
            df = load_processed_report(rec)
            if df is not None and not df.empty:
                frames.append(df.copy())
        except Exception:
            continue
    return pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()


def _first_nonblank(row: pd.Series, cols: list[str]) -> str:
    for col in cols:
        if col in row.index:
            val = _safe_text(row.get(col))
            if val:
                return _norm_id(val)
    return ""


def _build_lookup(client_name: str | None) -> dict:
    df = _load_lookup_df(client_name)
    lookup = {"campaign_adgroup": {}, "keyword": {}, "product_target": {}}
    if df.empty:
        return lookup
    for _, r in df.iterrows():
        campaign = _safe_text(r.get("campaign")) or _raw(r, "campaign name", "campaign name (informational only)")
        ad_group = _safe_text(r.get("ad_group")) or _raw(r, "ad group name", "ad group name (informational only)")
        cid = _safe_text(r.get("campaign_id")) or _raw(r, "campaign id")
        agid = _safe_text(r.get("ad_group_id")) or _raw(r, "ad group id")
        cid = _norm_id(cid); agid = _norm_id(agid)
        if campaign and ad_group and cid and agid:
            lookup["campaign_adgroup"][( _norm(campaign), _norm(ad_group) )] = (cid, agid)
        kw = _safe_text(r.get("keyword_text")) or _raw(r, "keyword text") or _safe_text(r.get("target"))
        mt = _safe_text(r.get("match_type")) or _raw(r, "match type")
        kid = _safe_text(r.get("keyword_id")) or _raw(r, "keyword id")
        kid = _norm_id(kid)
        if campaign and ad_group and kw and kid:
            lookup["keyword"][( _norm(campaign), _norm(ad_group), _norm(kw), _norm(mt) )] = kid
            lookup["keyword"][( _norm(campaign), _norm(ad_group), _norm(kw), "" )] = kid
        expr = _safe_text(r.get("product_targeting_expression")) or _safe_text(r.get("resolved_product_targeting_expression")) or _raw(r, "product targeting expression", "resolved product targeting expression", "resolved product targeting expression (informational only)") or _safe_text(r.get("target"))
        ptid = _safe_text(r.get("product_targeting_id")) or _raw(r, "product targeting id")
        ptid = _norm_id(ptid)
        if campaign and ad_group and expr and ptid:
            expr_norms = {_norm(expr), _norm(_product_expression_from_target(expr))}
            asin = _asin_from_text(expr)
            if asin:
                expr_norms.add(_norm(asin))
                expr_norms.add(_norm(f'asin="{asin.lower()}"'))
            for expr_norm in expr_norms:
                if expr_norm:
                    lookup["product_target"][( _norm(campaign), _norm(ad_group), expr_norm )] = ptid
    return lookup


def _fill_parent_ids(data: dict, row: pd.Series, lookup: dict) -> None:
    if data.get("Campaign ID") and data.get("Ad Group ID"):
        return
    campaign = data.get("Campaign Name") or _safe_text(row.get("campaign"))
    ad_group = data.get("Ad Group Name") or _safe_text(row.get("ad_group"))
    ids = lookup.get("campaign_adgroup", {}).get((_norm(campaign), _norm(ad_group)))
    if ids:
        data["Campaign ID"], data["Ad Group ID"] = ids


def _fill_keyword_id(data: dict, row: pd.Series, lookup: dict) -> None:
    if data.get("Keyword ID"):
        return
    campaign = data.get("Campaign Name") or _safe_text(row.get("campaign"))
    ad_group = data.get("Ad Group Name") or _safe_text(row.get("ad_group"))
    kw = data.get("Keyword Text") or _safe_text(row.get("target")) or _raw(row, "keyword text")
    mt = data.get("Match Type") or _raw(row, "match type")
    kid = lookup.get("keyword", {}).get((_norm(campaign), _norm(ad_group), _norm(kw), _norm(mt))) or lookup.get("keyword", {}).get((_norm(campaign), _norm(ad_group), _norm(kw), ""))
    if kid:
        data["Keyword ID"] = kid


def _fill_product_targeting_id(data: dict, row: pd.Series, lookup: dict) -> None:
    if data.get("Product Targeting ID"):
        return
    campaign = data.get("Campaign Name") or _safe_text(row.get("campaign"))
    ad_group = data.get("Ad Group Name") or _safe_text(row.get("ad_group"))
    expr = data.get("Product Targeting Expression") or _safe_text(row.get("target")) or _raw(row, "product targeting expression", "resolved product targeting expression", "resolved product targeting expression (informational only)")
    exprs = [_norm(expr), _norm(_product_expression_from_target(expr))]
    asin = _asin_from_text(expr)
    if asin:
        exprs += [_norm(asin), _norm(f'asin="{asin.lower()}"')]
    for expr_norm in exprs:
        ptid = lookup.get("product_target", {}).get((_norm(campaign), _norm(ad_group), expr_norm))
        if ptid:
            data["Product Targeting ID"] = ptid
            return


def _bid_update_row(row: pd.Series, ad_type: str, lookup: dict | None = None) -> dict:
    lookup = lookup or {"campaign_adgroup": {}, "keyword": {}, "product_target": {}}
    suggested_bid = _safe_float(row.get("suggested_bid"), 0)
    is_pt = _is_product_target(row)
    entity = "Product Targeting" if is_pt else "Keyword"
    keyword_text = "" if is_pt else (_raw(row, "keyword text") or _safe_text(row.get("keyword_text")) or _safe_text(row.get("target")))
    product_expr = ""
    if is_pt:
        product_expr = _raw(row, "product targeting expression", "resolved product targeting expression", "resolved product targeting expression (informational only)") or _safe_text(row.get("product_targeting_expression")) or _safe_text(row.get("target"))
        product_expr = _product_expression_from_target(product_expr)
    data = {
        "Product": _product(ad_type, row),
        "Entity": entity,
        "Operation": "Update",
        "Campaign ID": _norm_id(_safe_text(row.get("campaign_id")) or _raw(row, "campaign id")),
        "Ad Group ID": _norm_id(_safe_text(row.get("ad_group_id")) or _raw(row, "ad group id")),
        "Keyword ID": "" if is_pt else _norm_id(_safe_text(row.get("keyword_id")) or _raw(row, "keyword id")),
        "Product Targeting ID": _norm_id(_safe_text(row.get("product_targeting_id")) or _raw(row, "product targeting id")) if is_pt else "",
        "Campaign Name": _raw(row, "campaign name", "campaign name (informational only)") or _safe_text(row.get("campaign")),
        "Ad Group Name": _raw(row, "ad group name", "ad group name (informational only)") or _safe_text(row.get("ad_group")),
        "State": _raw(row, "state") or _safe_text(row.get("state")) or "enabled",
        "Keyword Text": keyword_text,
        "Match Type": "" if is_pt else (_raw(row, "match type") or _safe_text(row.get("match_type")) or "exact"),
        "Bid": round(suggested_bid, 2) if suggested_bid > 0 else "",
        "Budget": "",
        "Daily Budget": "",
        "Placement Type": "",
        "Placement %": "",
        "Portfolio ID": _norm_id(_safe_text(row.get("portfolio_id")) or _raw(row, "portfolio id")),
        "Product Targeting Expression": product_expr,
    }
    _fill_parent_ids(data, row, lookup)
    if is_pt:
        _fill_product_targeting_id(data, row, lookup)
    else:
        _fill_keyword_id(data, row, lookup)
    return data


def _negative_row(row: pd.Series, ad_type: str, lookup: dict | None = None) -> dict:
    lookup = lookup or {"campaign_adgroup": {}, "keyword": {}, "product_target": {}}
    term = _target_for_negative(row)
    is_asin = bool(_asin_from_text(term))
    data = {
        "Product": _product(ad_type, row),
        "Entity": "Negative Product Targeting" if is_asin else "Negative Keyword",
        "Operation": "Create",
        "Campaign ID": _norm_id(_safe_text(row.get("campaign_id")) or _raw(row, "campaign id")),
        "Ad Group ID": _norm_id(_safe_text(row.get("ad_group_id")) or _raw(row, "ad group id")),
        "Keyword ID": "",
        "Product Targeting ID": "",
        "Campaign Name": _raw(row, "campaign name", "campaign name (informational only)") or _safe_text(row.get("campaign")),
        "Ad Group Name": _raw(row, "ad group name", "ad group name (informational only)") or _safe_text(row.get("ad_group")),
        "State": "enabled",
        "Keyword Text": "" if is_asin else term,
        "Match Type": "" if is_asin else "negative exact",
        "Bid": "",
        "Budget": "",
        "Daily Budget": "",
        "Placement Type": "",
        "Placement %": "",
        "Portfolio ID": _norm_id(_safe_text(row.get("portfolio_id")) or _raw(row, "portfolio id")),
        "Product Targeting Expression": _product_expression_from_target(term) if is_asin else "",
    }
    _fill_parent_ids(data, row, lookup)
    return data


def _campaign_budget_row(row: pd.Series, ad_type: str, lookup: dict | None = None) -> dict:
    lookup = lookup or {"campaign_adgroup": {}, "keyword": {}, "product_target": {}}
    budget = _safe_float(row.get("suggested_budget"), 0) or _safe_float(row.get("daily_budget"), 0)
    data = {
        "Product": _product(ad_type, row),
        "Entity": "Campaign",
        "Operation": "Update",
        "Campaign ID": _norm_id(_safe_text(row.get("campaign_id")) or _raw(row, "campaign id")),
        "Ad Group ID": "",
        "Keyword ID": "",
        "Product Targeting ID": "",
        "Campaign Name": _raw(row, "campaign name", "campaign name (informational only)") or _safe_text(row.get("campaign")),
        "Ad Group Name": "",
        "State": _raw(row, "state") or _safe_text(row.get("state")) or "enabled",
        "Keyword Text": "",
        "Match Type": "",
        "Bid": "",
        "Budget": "",
        "Daily Budget": round(budget, 2) if budget > 0 else "",
        "Placement Type": "",
        "Placement %": "",
        "Portfolio ID": _norm_id(_safe_text(row.get("portfolio_id")) or _raw(row, "portfolio id")),
        "Product Targeting Expression": "",
    }
    return data


def _truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "include"}


def _brand_safety_uploadable(row: pd.Series) -> bool:
    if not _truthy(row.get("include_in_bulk_upload", False)):
        return False
    campaign = _safe_text(row.get("campaign"))
    ad_group = _safe_text(row.get("ad_group"))
    return bool(campaign and ad_group and campaign != "not available" and ad_group != "not available")


def _is_valid_upload_row(row: dict) -> tuple[bool, str]:
    entity = _safe_text(row.get("Entity")).lower()
    op = _safe_text(row.get("Operation")).lower()
    if op == "update" and entity == "keyword":
        if not row.get("Campaign ID") or not row.get("Ad Group ID") or not row.get("Keyword ID"):
            return False, "Keyword update missing Campaign ID, Ad Group ID, or Keyword ID"
        if row.get("Bid") in {"", None}:
            return False, "Keyword update missing Bid"
    if op == "update" and entity == "product targeting":
        if not row.get("Campaign ID") or not row.get("Ad Group ID") or not row.get("Product Targeting ID"):
            return False, "Product targeting update missing Campaign ID, Ad Group ID, or Product Targeting ID"
        if row.get("Bid") in {"", None}:
            return False, "Product targeting update missing Bid"
    if op == "create" and entity in {"negative keyword", "negative product targeting"}:
        if not row.get("Campaign ID") or not row.get("Ad Group ID"):
            return False, "Negative create missing Campaign ID or Ad Group ID"
        if entity == "negative keyword" and (not row.get("Keyword Text") or not row.get("Match Type")):
            return False, "Negative keyword missing Keyword Text or Match Type"
        if entity == "negative product targeting" and not row.get("Product Targeting Expression"):
            return False, "Negative product targeting missing Product Targeting Expression"
    if op == "update" and entity == "campaign":
        if not row.get("Campaign ID") or row.get("Daily Budget") in {"", None}:
            return False, "Campaign update missing Campaign ID or Daily Budget"
    return True, ""


def build_bulk_rows(actions: pd.DataFrame, ad_type: str, client_name: str | None = None, include_invalid: bool = False) -> pd.DataFrame:
    headers = _template_headers()
    rows: list[dict] = []
    invalid: list[dict] = []
    lookup = _build_lookup(client_name)
    if actions is None or actions.empty:
        return pd.DataFrame(columns=headers)
    for _, row in actions.iterrows():
        reason = _safe_text(row.get("reason_code"))
        category = _safe_text(row.get("category"))
        suggested_bid = _safe_float(row.get("suggested_bid"), 0)
        candidates: list[dict] = []
        if reason == "high_click_no_order":
            candidates.append(_negative_row(row, ad_type, lookup))
        elif reason.startswith(BRAND_REASON_PREFIX):
            if _brand_safety_uploadable(row):
                candidates.append(_negative_row(row, ad_type, lookup))
        if suggested_bid > 0 and not reason.startswith(BRAND_REASON_PREFIX):
            candidates.append(_bid_update_row(row, ad_type, lookup))
        if category.lower().startswith("budget"):
            candidates.append(_campaign_budget_row(row, ad_type, lookup))
        for c in candidates:
            valid, msg = _is_valid_upload_row(c)
            if valid:
                rows.append(c)
            else:
                c["_validation_error"] = msg
                c["_source_reason_code"] = reason
                c["_source_target"] = _safe_text(row.get("target"))
                invalid.append(c)
    out = pd.DataFrame(rows)
    if out.empty:
        out = pd.DataFrame(columns=headers)
    else:
        out = out.drop_duplicates()
        for h in headers:
            if h not in out.columns:
                out[h] = ""
        out = out[headers]
    if include_invalid and invalid:
        bad = pd.DataFrame(invalid)
        for h in headers:
            if h not in bad.columns:
                bad[h] = ""
        return out, bad[headers + ["_validation_error", "_source_reason_code", "_source_target"]]
    return out


def build_bulk_validation(actions: pd.DataFrame, ad_type: str, client_name: str | None = None) -> pd.DataFrame:
    rows, invalid = build_bulk_rows(actions, ad_type, client_name=client_name, include_invalid=True)
    return invalid


def _write_template_output(path: Path, rows: pd.DataFrame, skipped: pd.DataFrame | None = None) -> None:
    if load_workbook and TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb["Output"] if "Output" in wb.sheetnames else wb[wb.sheetnames[0]]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        headers = [_safe_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h]
        for r_idx, record in enumerate(rows.to_dict(orient="records"), start=2):
            for c_idx, header in enumerate(headers, start=1):
                ws.cell(row=r_idx, column=c_idx).value = record.get(header, "")
        # Keep uploads clean: only Output is required. Add skipped rows only if no Amazon upload will see the file.
        wb.save(path)
    else:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            rows.to_excel(writer, index=False, sheet_name="Output")


def export_action_review(client_name: str, actions: pd.DataFrame) -> Path:
    path = _out_dir(client_name) / f"{slugify(client_name)}_action_review_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        actions.to_excel(writer, index=False, sheet_name="Approved_Actions")
    return path


def export_bulk_for_ad_type(client_name: str, ad_type: str, actions: pd.DataFrame) -> Path:
    ad_type = ad_type.upper()
    rows = build_bulk_rows(actions, ad_type, client_name=client_name)
    path = _out_dir(client_name) / f"{slugify(client_name)}_{ad_type}_bulk_upload_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    _write_template_output(path, rows)
    return path


def export_bulk_review_workbook(client_name: str, actions_by_ad_type: dict[str, pd.DataFrame], audit: dict | None = None) -> Path:
    path = _out_dir(client_name) / f"{slugify(client_name)}_bulk_audit_review_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if audit:
            pd.DataFrame([{"summary": audit.get("summary", ""), "mode": audit.get("mode", "")}]).to_excel(writer, index=False, sheet_name="AI_Audit")
            pd.DataFrame({"warnings": audit.get("warnings", [])}).to_excel(writer, index=False, sheet_name="Warnings")
        for ad_type, actions in actions_by_ad_type.items():
            actions = actions if actions is not None else pd.DataFrame()
            actions.to_excel(writer, index=False, sheet_name=f"{ad_type}_Approved_Actions"[:31])
            build_bulk_rows(actions, ad_type, client_name=client_name).to_excel(writer, index=False, sheet_name=f"{ad_type}_Upload_Preview"[:31])
            build_bulk_validation(actions, ad_type, client_name=client_name).to_excel(writer, index=False, sheet_name=f"{ad_type}_Skipped"[:31])
    return path


def zip_bulk_files(paths: list[Path], client_name: str) -> Path:
    zip_path = _out_dir(client_name) / f"{slugify(client_name)}_split_bulk_uploads_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for path in paths:
            if path and Path(path).exists():
                z.write(path, arcname=Path(path).name)
    return zip_path


def export_bulk_prep(client_name: str, actions: pd.DataFrame) -> Path:
    grouped = {}
    if actions is not None and not actions.empty and "ad_type" in actions.columns:
        for ad_type, df in actions.groupby(actions["ad_type"].astype(str).str.upper()):
            grouped[ad_type] = df.copy()
    else:
        grouped["SP"] = actions
    return export_bulk_review_workbook(client_name, grouped)
